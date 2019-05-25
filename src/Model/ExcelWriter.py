# -*- coding: utf-8 -*-

from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.chart import (
    LineChart,
    Reference,
)
import sys


def write_table(workbook, table):
    # create the sheet
    worksheet = workbook.create_sheet('Table', 0)
    worksheet.sheet_view.rightToLeft = True
    # split to rows by "<"
    table_info = table.split("<")[1:]
    titles = table_info[0][:-1]
    titles = titles.split("!")
    # add the titles to the table
    for i, t in enumerate(titles):
        worksheet.cell(row=1, column=i + 1, value=t)

    table_info = table_info[1:]
    # go over the rows of information
    for row in range(len(table_info)):
        # split the row data by "!"
        temp_row = table_info[row].split("!")
        for i, t in enumerate(temp_row):
            worksheet.cell(row=2 + row, column=i + 1, value=t)

    return workbook


def write_line_chart(workbook, info, chart_name):
    # create the sheet
    worksheet = workbook.create_sheet(chart_name)
    worksheet.sheet_view.rightToLeft = True
    # split to rows by "<"
    chart_info = info.split("<")
    chart_lines_names = []
    # get the line names
    for i in chart_info:
        line_name = i.split("!")[0]
        chart_lines_names.append(line_name)
    # create the rows
    rows = []
    columns = []
    rows.append(chart_lines_names)
    for info in chart_info:
        # remove unnecessary chars in the row
        info = info.replace("Data[", "").replace(",null]", "")
        info = info.split("!")
        # remove the name
        info.pop(0)
        col = []
        # each value is game number and avg reaction time
        for value in info:
            v = value.split(",")[-1]
            col.append(v)
        columns.append(col)

    # convert the columns to rows
    for i in range(len(max(columns))):
        row = []
        for c in columns:
            if i >= len(c):
                row.append('')
            else:
                row.append(float(c[i])) if '.' in c[i] else row.append('')

        rows.append(row)

    # add the rows to the sheet
    for row in rows:
        row.pop(0)
        worksheet.append(row)

    c1 = LineChart()
    c1.title = chart_name
    c1.y_axis.title = "Reaction Time"
    c1.x_axis.title = 'Game Number'
    c1.x_axis.tickLblSkip = 1
    c1.x_axis.tickMarkSkip = 1

    data = Reference(worksheet, min_col=1, min_row=1, max_col=len(chart_lines_names), max_row=len(rows))
    c1.add_data(data, titles_from_data=True)

    style_symbols = ["triangle", "diamond", "square", "x"]
    for i in range(len(c1.series)):
        # Style the lines
        s1 = c1.series[i]
        s1.marker.symbol = style_symbols[i]
    worksheet.add_chart(c1, "A1")

    return workbook


def main():
    charts = []
    # read all arguments from the java program - table and charts
    for i in range(len(sys.argv)):
        if i == 0:
            continue
        charts.append(sys.argv[i])
    # set file path
    file_path = charts.pop(0)
    table = charts.pop(0)
    shapes = charts.pop(0)
    textures = charts.pop(0)

    workbook = Workbook()
    workbook.remove(workbook['Sheet'])
    workbook.encoding = "utf-8"
    workbook = write_table(workbook, table)
    workbook = write_line_chart(workbook, shapes, "Shapes Chart")
    workbook = write_line_chart(workbook, textures, "Textures Chart")

    # save workbook
    try:
        workbook.save(file_path)
        workbook.close()
        print("file saved!")
    except:
        print("file not saved")
        print("Unexpected error:", sys.exc_info()[0])
        raise

main()
